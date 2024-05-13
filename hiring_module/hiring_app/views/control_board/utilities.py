from functools import wraps
from django.shortcuts import redirect

from hiring_app.model import CustomUser
from django.contrib.auth.models import Group

from hiring_app.model.cex_contract_request_model import CEXContractRequest
from hiring_app.model.monitoring_contract_request_model import MonitoringContractRequest
from hiring_app.model.provision_of_services_request_model import ProvisionOfServicesContractRequest
from django.http import HttpResponse
from openpyxl import Workbook

from datetime import datetime
from django.db.models import Q
from django.utils import timezone
from itertools import chain


# Decorator to redirect users to the correct dashboard based on their role
def role_redirect(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Check if the user is already on the correct dashboard
        if request.user.groups.filter(name='admin').exists() and request.path != '/hiring_app/administrator_dashboard/':
            return redirect('hiring_app:administrator_dashboard')
        elif request.user.groups.filter(name='leader').exists() and request.path != '/hiring_app/leader_dashboard/':
            return redirect('hiring_app:leader_dashboard')
        elif request.user.groups.filter(name='manager').exists() and request.path != '/hiring_app/manager_dashboard/':
            return redirect('hiring_app:manager_dashboard')
        elif not request.user.groups.exists() and request.path != '/hiring_app/external_user_dashboard/':
            return redirect('hiring_app:external_user_dashboard')
        
        # Call the original view function if no redirection is needed
        return view_func(request, *args, **kwargs)
    
    return wrapper

# Decorator to ensure that only users with the 'admin' role can access a view
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.groups.filter(name='admin').exists():
            if request.user.groups.filter(name='leader').exists():
                return redirect('hiring_app:leader_dashboard')
            elif request.user.groups.filter(name='manager').exists():
                return redirect('hiring_app:manager_dashboard')
            else:
                return redirect('hiring_app:external_user_dashboard')
        return view_func(request, *args, **kwargs)
    
    return wrapper

# Decorator to ensure that only users with the 'leader' role can access a view
def leader_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.groups.filter(name='leader').exists():
            if request.user.groups.filter(name='admin').exists():
                return redirect('hiring_app:administrator_dashboard')
            elif request.user.groups.filter(name='manager').exists():
                return redirect('hiring_app:manager_dashboard')
            else:
                return redirect('hiring_app:external_user_dashboard')
        return view_func(request, *args, **kwargs)
    
    return wrapper

# Decorator to ensure that only users with the 'manager' role can access a view
def manager_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.groups.filter(name='manager').exists():
            if request.user.groups.filter(name='admin').exists():
                return redirect('hiring_app:administrator_dashboard')
            elif request.user.groups.filter(name='leader').exists():
                return redirect('hiring_app:leader_dashboard')
            else:
                return redirect('hiring_app:external_user_dashboard')
        return view_func(request, *args, **kwargs)
    
    return wrapper
def get_requests(user):
    groups = [group.name for group in user.groups.all()]
    requests_CEX = CEXContractRequest.objects.none()
    requests_monitoring = MonitoringContractRequest.objects.none()
    requests_pos = ProvisionOfServicesContractRequest.objects.none()

    if 'admin' in groups:
        requests_CEX = CEXContractRequest.objects.all()
        requests_monitoring = MonitoringContractRequest.objects.all()
        requests_pos = ProvisionOfServicesContractRequest.objects.all()
    elif 'leader' in groups:
        requests_CEX = CEXContractRequest.objects.filter(leader_assigned_to=user.id)
        requests_monitoring = MonitoringContractRequest.objects.filter(leader_assigned_to=user.id)
        requests_pos = ProvisionOfServicesContractRequest.objects.filter(leader_assigned_to=user.id)
    elif 'manager' in groups:
        requests_CEX = CEXContractRequest.objects.filter(manager_assigned_to=user.id)
        requests_monitoring = MonitoringContractRequest.objects.filter(manager_assigned_to=user.id)
        requests_pos = ProvisionOfServicesContractRequest.objects.filter(manager_assigned_to=user.id)
      
    for request in requests_CEX:
        request.request_type = 'CEX'

    for request in requests_monitoring:
        request.request_type = 'Monitoría'

    for request in requests_pos:
        request.request_type = 'Honorarios' 
          
    groupManager = Group.objects.get(name='manager')
    groupLeader = Group.objects.get(name='leader')       
    managers = list(CustomUser.objects.filter(groups=groupManager))
    leaders = list(CustomUser.objects.filter(groups=groupLeader))
    
    current_date = timezone.now()
    current_month = current_date.month
    current_year = current_date.year

    requests = list(requests_CEX) + list(requests_monitoring) + list(requests_pos)

    requests_month = list(
        chain(
            requests_CEX.filter(start_date__month=current_month, start_date__year=current_year),
            requests_monitoring.filter(start_date__month=current_month, start_date__year=current_year),
            requests_pos.filter(start_date__month=current_month, start_date__year=current_year)
        )
    )

    filled_requests = list(
        chain(
            requests_CEX.filter(start_date__month=current_month, start_date__year=current_year, state='filed'),
            requests_monitoring.filter(start_date__month=current_month, start_date__year=current_year, state='filed'),
            requests_pos.filter(start_date__month=current_month, start_date__year=current_year, state='filed')
        )
    )

    reviewed_requests = list(
        chain(
            requests_CEX.filter(start_date__month=current_month, start_date__year=current_year, state='review'),
            requests_monitoring.filter(start_date__month=current_month, start_date__year=current_year, state='review'),
            requests_pos.filter(start_date__month=current_month, start_date__year=current_year, state='review')
        )
    )
    
    
    return {
        'requests': requests,
        'requests_month': requests_month,
        'filled_requests': filled_requests,
        'reviewed_requests': reviewed_requests,
        'for_validate_requests':  list(requests_CEX.filter(state__in=['pending', 'incomplete'])) + list(requests_monitoring.filter(state__in=['pending', 'incomplete'])) + list(requests_pos.filter(state__in=['pending', 'incomplete'])),
        'leaders': leaders,
        'managers': managers,
    }

        
        
def export_requests(user):
    groups = [group.name for group in user.groups.all()]
    requests_CEX = CEXContractRequest.objects.none()
    requests_monitoring = MonitoringContractRequest.objects.none()
    requests_pos = ProvisionOfServicesContractRequest.objects.none()

    if 'admin' in groups:
        requests_CEX = CEXContractRequest.objects.all()
        requests_monitoring = MonitoringContractRequest.objects.all()
        requests_pos = ProvisionOfServicesContractRequest.objects.all()
    elif 'leader' in groups:
        requests_CEX = CEXContractRequest.objects.filter(leader_assigned_to=user.id)
        requests_monitoring = MonitoringContractRequest.objects.filter(leader_assigned_to=user.id)
        requests_pos = ProvisionOfServicesContractRequest.objects.filter(leader_assigned_to=user.id)
    elif 'manager' in groups:
        requests_CEX = CEXContractRequest.objects.filter(manager_assigned_to=user.id)
        requests_monitoring = MonitoringContractRequest.objects.filter(manager_assigned_to=user.id)
        requests_pos = ProvisionOfServicesContractRequest.objects.filter(manager_assigned_to=user.id)

    

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Solicitudes de Contratación"
    
    headers = [
    "ID",
    "Start Date",
    "Completion Date",
    "Estimated Completion Date",
    "Current State Start",
    "State",
    "Manager Assigned To",
    "Leader Assigned To",
    "Created By",
    ]
    
    ws1.append(headers)
    
    for request in (list(requests_CEX) + list(requests_monitoring) + list(requests_pos)):
        row_data = [
            str(request.id),
            request.start_date.strftime('%Y-%m-%d'),
            request.completion_date.strftime('%Y-%m-%d %H:%M:%S') if request.completion_date else None,
            request.estimated_completion_date.strftime('%Y-%m-%d') if request.estimated_completion_date else None,
            request.current_state_start.strftime('%Y-%m-%d %H:%M:%S'),
            request.state,
            f'{request.manager_assigned_to.first_name} {request.manager_assigned_to.last_name}' if request.manager_assigned_to else 'sin asignar',
            f'{request.leader_assigned_to.first_name} {request.leader_assigned_to.last_name}' if request.leader_assigned_to else 'sin asignar',
            f'{request.created_by.first_name} {request.created_by.last_name}' if request.created_by else 'sin asignar',
        ]
        ws1.append(row_data)
    
    
    ws2 = wb.create_sheet(title="CEX")
    
    headers = [
    "ID",
    "Start Date",
    "Completion Date",
    "Estimated Completion Date",
    "Current State Start",
    "State",
    "Manager Assigned To",
    "Leader Assigned To",
    "Created By",
    "Hiree Full Name",
    "Hiree ID",
    "Hiree Cellphone",
    "Hiree Email",
    "Cenco",
    "Request Motive",
    "Banking Entity",
    "Bank Account Type",
    "Bank Account Number",
    "EPS",
    "Pension Fund",
    "ARL",
    "Contract Value",
    "Charge Account",
    "RUT"
    ]
    ws2.append(headers)
    
    for cex_request in requests_CEX:
        row_data = [
            str(cex_request.id),
            cex_request.start_date.strftime('%Y-%m-%d'),
            cex_request.completion_date.strftime('%Y-%m-%d %H:%M:%S') if cex_request.completion_date else None,
            cex_request.estimated_completion_date.strftime('%Y-%m-%d') if cex_request.estimated_completion_date else None,
            cex_request.current_state_start.strftime('%Y-%m-%d %H:%M:%S'),
            cex_request.state,
            f'{cex_request.manager_assigned_to.first_name} {cex_request.manager_assigned_to.last_name}' if cex_request.manager_assigned_to else 'sin asignar',
            f'{cex_request.leader_assigned_to.first_name} {cex_request.leader_assigned_to.last_name}' if cex_request.leader_assigned_to else 'sin asignar',
            f'{cex_request.created_by.first_name} {cex_request.created_by.last_name}' if cex_request.created_by else 'sin asignar',
            cex_request.hiree_full_name,
            cex_request.hiree_id,
            cex_request.hiree_cellphone,
            cex_request.hiree_email,
            cex_request.cenco,
            cex_request.request_motive,
            cex_request.banking_entity,
            cex_request.bank_account_type,
            cex_request.bank_account_number,
            cex_request.eps,
            cex_request.pension_fund,
            cex_request.arl,
            cex_request.contract_value,
            cex_request.charge_account,
            cex_request.rut.url if cex_request.rut else ""
        ]
        ws2.append(row_data)

    ws3 = wb.create_sheet(title="Monitorías")
    
    headers = [
    "ID",
    "Start Date",
    "Completion Date",
    "Estimated Completion Date",
    "Current State Start",
    "State",
    "Manager Assigned To",
    "Leader Assigned To",
    "Created By",
    "Cenco",
    "Has Money in Cenco",
    "Cenco Manager",
    "Monitoring Type",
    "Student Code",
    "Student Full Name",
    "Student ID",
    "Student Email",
    "Student Cellphone",
    "Daviplata",
    "Course or Project",
    "Monitoring Description",
    "Weekly Hours",
    "Total Value to Pay",
    "Is Unique Payment"
    ]
    ws3.append(headers)

    for monitoring_request in requests_monitoring:
        row_data = [
            str(monitoring_request.id),
            monitoring_request.start_date.strftime('%Y-%m-%d'),
            monitoring_request.completion_date.strftime('%Y-%m-%d %H:%M:%S') if monitoring_request.completion_date else None,
            monitoring_request.estimated_completion_date.strftime('%Y-%m-%d') if monitoring_request.estimated_completion_date else None,
            monitoring_request.current_state_start.strftime('%Y-%m-%d %H:%M:%S'),
            monitoring_request.state,
            f'{monitoring_request.manager_assigned_to.first_name} {monitoring_request.manager_assigned_to.last_name}' if monitoring_request.manager_assigned_to else 'sin asignar',
            f'{monitoring_request.leader_assigned_to.first_name} {monitoring_request.leader_assigned_to.last_name}' if monitoring_request.leader_assigned_to else 'sin asignar',
            f'{monitoring_request.created_by.first_name} {monitoring_request.created_by.last_name}' if monitoring_request.created_by else 'sin asignar',
            monitoring_request.cenco,
            monitoring_request.has_money_in_cenco,
            monitoring_request.cenco_manager,
            monitoring_request.monitoring_type,
            monitoring_request.student_code,
            monitoring_request.student_full_name,
            monitoring_request.student_id,
            monitoring_request.student_email,
            monitoring_request.student_cellphone,
            monitoring_request.daviplata,
            monitoring_request.course_or_proyect,
            monitoring_request.monitoring_description,
            monitoring_request.weekly_hours,
            monitoring_request.total_value_to_pay,
            monitoring_request.is_unique_payment
        ]
        ws3.append(row_data)

    ws4 = wb.create_sheet(title="Honorarios")
    
    headers = [
    "ID",
    "Start Date",
    "Completion Date",
    "Estimated Completion Date",
    "Current State Start",
    "State",
    "Manager Assigned To",
    "Leader Assigned To",
    "Created By",
    "Hiree Full Name",
    "Hiree ID",
    "Hiree Cellphone",
    "Hiree Email",
    "Cenco",
    "Request Motive",
    "Banking Entity",
    "Bank Account Type",
    "Bank Account Number",
    "EPS",
    "Pension Fund",
    "ARL",
    "Contract Value",
    "Charge Account",
    "RUT",
    "Course Name",
    "Period",
    "Group",
    "Intensity",
    "Total Hours",
    "Course Code",
    "Students Quantity",
    "Additional Hours"
    ]

    ws4.append(headers)
    
    for post_request in requests_pos:
        row_data = [
            str(post_request.id),
            post_request.start_date.strftime('%Y-%m-%d'),
            post_request.completion_date.strftime('%Y-%m-%d %H:%M:%S') if post_request.completion_date else None,
            post_request.estimated_completion_date.strftime('%Y-%m-%d') if post_request.estimated_completion_date else None,
            post_request.current_state_start.strftime('%Y-%m-%d %H:%M:%S'),
            post_request.state,
            f'{post_request.manager_assigned_to.first_name} {post_request.manager_assigned_to.last_name}' if post_request.manager_assigned_to else 'sin asignar',
            f'{post_request.leader_assigned_to.first_name} {post_request.leader_assigned_to.last_name}' if post_request.leader_assigned_to else 'sin asignar',
            f'{post_request.created_by.first_name} {post_request.created_by.last_name}' if post_request.created_by else 'sin asignar',
            post_request.hiree_full_name,
            post_request.hiree_id,
            post_request.hiree_cellphone,
            post_request.hiree_email,
            post_request.cenco,
            post_request.request_motive,
            post_request.banking_entity,
            post_request.bank_account_type,
            post_request.bank_account_number,
            post_request.eps,
            post_request.pension_fund,
            post_request.arl,
            post_request.contract_value,
            post_request.charge_account,
            post_request.rut.url if post_request.rut else "", 
            post_request.course_name,
            post_request.period,
            post_request.group,
            post_request.intensity,
            post_request.total_hours,
            post_request.course_code,
            post_request.students_quantity,
            post_request.additional_hours
        ]
        ws4.append(row_data)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="solicitudes.xlsx"'
    wb.save(response)

    return response